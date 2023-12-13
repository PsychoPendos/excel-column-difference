import openpyxl

wb = openpyxl.load_workbook('D:\FP\Downloads\sheet.xlsx')
ws = wb['List1']
list1 = ws['A']
list2 = ws['B']
mylist1 = [cell.value for cell in list1]
mylist2 = [cell.value for cell in list2]

missing = []
notmissing = []

for itemin1 in mylist1:
    if itemin1 not in mylist2:
        missing.append(itemin1)
    else:
        notmissing.append(itemin1)

wb.create_sheet('missing')
wb.create_sheet('notmissing')

ws = wb['missing']

for i, item in enumerate(missing):
    ws.cell(row=i+1, column=1, value=item)
    wb.save('D:\FP\Downloads\sheet.xlsx')

ws = wb['notmissing']

for i, item in enumerate(notmissing):
    ws.cell(row=i+1, column=1, value=item)
    wb.save('D:\FP\Downloads\sheet.xlsx')
